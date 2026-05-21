from flask import Flask, request, jsonify, send_file, render_template_string
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from collections import defaultdict
import os, io, uuid

app = Flask(__name__)
UPLOAD_FOLDER  = os.path.join(os.path.dirname(__file__), 'uploads')
RESULT_FOLDER  = os.path.join(os.path.dirname(__file__), 'resultados')
BANCOS_IGNORAR_PREFIXOS = ('prata', 'ph', 'presença', 'presenca')

# ── helpers ────────────────────────────────────────────────
def clean_cpf(v):
    if not v: return None
    return str(v).replace('.','').replace('-','').replace(' ','').strip().replace('\xa0','')

def clean_prop(v):
    if not v: return None
    return str(v).strip()

def banco_ignorar(v):
    if not v: return False
    b = str(v).strip().lower()
    return any(b.startswith(p) for p in BANCOS_IGNORAR_PREFIXOS)

def soma(lista):
    total = 0.0
    for item in lista:
        row = item[0] if (isinstance(item, tuple) and isinstance(item[0], tuple)) else item
        v = row[5]
        if isinstance(v, (int, float)): total += v
    return total

def por_banco(lista):
    d = defaultdict(lambda: {'qtd': 0, 'val': 0.0})
    for item in lista:
        row = item[0] if (isinstance(item, tuple) and isinstance(item[0], tuple)) else item
        banco = str(row[12]).strip() if row[12] else 'SEM BANCO'
        d[banco]['qtd'] += 1
        if isinstance(row[5], (int, float)): d[banco]['val'] += row[5]
    return d

def processar(path_japa, path_nossa):
    wb1 = openpyxl.load_workbook(path_japa)
    ws1 = wb1.active
    wb2 = openpyxl.load_workbook(path_nossa)
    ws2 = wb2.active

    nossa_por_cpf  = {}
    nossa_por_prop = {}
    for row in ws2.iter_rows(min_row=4, values_only=True):
        if not any(row): continue
        cpf  = clean_cpf(row[20])
        prop = clean_prop(row[4])
        if cpf:  nossa_por_cpf[cpf]   = row
        if prop: nossa_por_prop[prop] = row

    pago_bate, pago_falta, npago_bate, npago_nao_bate = [], [], [], []

    for row in ws1.iter_rows(min_row=2, values_only=True):
        if row[1] in (None, 'CPF'): continue
        if banco_ignorar(row[12]):  continue
        status   = row[11]
        cpf      = clean_cpf(row[1])
        proposta = clean_prop(row[4])
        bate     = cpf in nossa_por_cpf or proposta in nossa_por_prop

        if status == 'PAGO':
            if bate:
                nr = nossa_por_cpf.get(cpf) or nossa_por_prop.get(proposta)
                pago_bate.append((row, nr))
            else:
                pago_falta.append(row)
        else:
            if bate:
                nr = nossa_por_cpf.get(cpf) or nossa_por_prop.get(proposta)
                npago_bate.append((row, nr))
            else:
                npago_nao_bate.append(row)

    v1, v2, v3, v4 = soma(pago_bate), soma(pago_falta), soma(npago_bate), soma(npago_nao_bate)
    bd_falta = por_banco(pago_falta)

    return {
        'pago_bate':        {'qtd': len(pago_bate),        'val': round(v1, 2)},
        'pago_falta':       {'qtd': len(pago_falta),       'val': round(v2, 2)},
        'npago_bate':       {'qtd': len(npago_bate),       'val': round(v3, 2)},
        'npago_nao_bate':   {'qtd': len(npago_nao_bate),   'val': round(v4, 2)},
        'total_qtd':        len(pago_bate)+len(pago_falta)+len(npago_bate)+len(npago_nao_bate),
        'total_val':        round(v1+v2+v3+v4, 2),
        'por_banco':        {k: v for k, v in sorted(bd_falta.items(), key=lambda x: -x[1]['val'])},
        '_listas':          (pago_bate, pago_falta, npago_bate, npago_nao_bate),
        '_header':          [cell.value for cell in ws1[1]],
    }

def gerar_xlsx(resultado):
    pago_bate, pago_falta, npago_bate, npago_nao_bate = resultado['_listas']
    header = resultado['_header']

    GREEN  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    RED    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    ORANGE = PatternFill(start_color="F4B942", end_color="F4B942", fill_type="solid")
    HFILL  = PatternFill(start_color="2F4F9F", end_color="2F4F9F", fill_type="solid")
    HFILL2 = PatternFill(start_color="9C0006", end_color="9C0006", fill_type="solid")
    HFONT  = Font(color="FFFFFF", bold=True)
    BOLD   = Font(bold=True)

    def write_header(ws, cols, fill=None):
        for i, v in enumerate(cols, 1):
            c = ws.cell(1, i, v)
            c.font = HFONT
            c.fill = fill or HFILL
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[1].height = 30

    def auto_width(ws):
        for col in ws.columns:
            ml = 0
            cl = col[0].column_letter
            for cell in col:
                try:
                    if cell.value: ml = max(ml, len(str(cell.value)))
                except: pass
            ws.column_dimensions[cl].width = min(ml + 4, 40)

    wb = openpyxl.Workbook()

    # ── Aba 1: Rascunho geral ──
    ws_r = wb.active
    ws_r.title = "Rascunho Geral"
    write_header(ws_r, header + ['SITUACAO NA NOSSA'])
    rn = 2
    for item in pago_bate:
        row = item[0]
        for i, v in enumerate(row, 1): ws_r.cell(rn, i, v).fill = GREEN
        ws_r.cell(rn, 6).number_format = 'R$ #,##0.00'
        ws_r.cell(rn, 16, 'PAGO - BATE COM NOSSA').fill = GREEN
        rn += 1
    for row in pago_falta:
        for i, v in enumerate(row, 1): ws_r.cell(rn, i, v).fill = YELLOW
        ws_r.cell(rn, 6).number_format = 'R$ #,##0.00'
        ws_r.cell(rn, 16, 'PAGO - FALTA NA NOSSA').fill = YELLOW
        rn += 1
    for item in npago_bate:
        row = item[0]
        for i, v in enumerate(row, 1): ws_r.cell(rn, i, v).fill = ORANGE
        ws_r.cell(rn, 6).number_format = 'R$ #,##0.00'
        ws_r.cell(rn, 16, f'NAO PAGO - TEM NA NOSSA - {row[11]}').fill = ORANGE
        rn += 1
    for row in npago_nao_bate:
        for i, v in enumerate(row, 1): ws_r.cell(rn, i, v).fill = RED
        ws_r.cell(rn, 6).number_format = 'R$ #,##0.00'
        ws_r.cell(rn, 16, f'NAO PAGO - SEM REGISTRO - {row[11]}').fill = RED
        rn += 1
    auto_width(ws_r)

    # ── Aba 2: Resumo ──
    ws_res = wb.create_sheet("Resumo")
    ws_res.cell(1, 1, "RESUMO DO BATIMENTO").font = Font(bold=True, size=14)
    for i, h in enumerate(["DESCRICAO", "QUANTIDADE", "VALOR LIBERADO (R$)"], 1):
        c = ws_res.cell(3, i, h); c.font = HFONT; c.fill = HFILL
    dados = [
        ("Pagos que BATEM com NOSSA",       len(pago_bate),       soma(pago_bate),       "006100", GREEN),
        ("Pagos que FALTAM na NOSSA",        len(pago_falta),      soma(pago_falta),      "9C5700", YELLOW),
        ("NAO PAGOS com registro na NOSSA",  len(npago_bate),      soma(npago_bate),      "9C0006", ORANGE),
        ("NAO PAGOS sem registro na NOSSA",  len(npago_nao_bate),  soma(npago_nao_bate),  "9C0006", RED),
    ]
    for i, (desc, qtd, val, cor, fill) in enumerate(dados, 4):
        ws_res.cell(i, 1, desc).fill = fill
        ws_res.cell(i, 2, qtd).font = Font(bold=True, color=cor)
        c = ws_res.cell(i, 3, val); c.font = Font(bold=True, color=cor); c.number_format = 'R$ #,##0.00'
    total_qtd = sum(len(l) for l in [pago_bate, pago_falta, npago_bate, npago_nao_bate])
    total_val = soma(pago_bate)+soma(pago_falta)+soma(npago_bate)+soma(npago_nao_bate)
    ws_res.cell(9, 1, "TOTAL GERAL").font = BOLD
    ws_res.cell(9, 2, total_qtd).font = BOLD
    c = ws_res.cell(9, 3, total_val); c.font = BOLD; c.number_format = 'R$ #,##0.00'
    ws_res.column_dimensions['A'].width = 42
    ws_res.column_dimensions['B'].width = 14
    ws_res.column_dimensions['C'].width = 22

    # ── Aba 3: Faltando na NOSSA ──
    ws_f = wb.create_sheet("Faltando na NOSSA")
    write_header(ws_f, header, HFILL2)
    for rn2, row in enumerate(pago_falta, 2):
        for i, v in enumerate(row, 1): ws_f.cell(rn2, i, v).fill = RED
        ws_f.cell(rn2, 6).number_format = 'R$ #,##0.00'
    tr = len(pago_falta) + 3
    ws_f.cell(tr, 2, "TOTAL").font = BOLD
    ws_f.cell(tr, 3, f'{len(pago_falta)} registros').font = BOLD
    c = ws_f.cell(tr, 6, soma(pago_falta)); c.font = BOLD; c.number_format = 'R$ #,##0.00'
    auto_width(ws_f)

    # ── Aba 4: Por Banco ──
    ws_b = wb.create_sheet("Faltando - Por Banco")
    ws_b.cell(1, 1, "PAGOS DA JAPA QUE FALTAM NA NOSSA - POR BANCO").font = Font(bold=True, size=12)
    ws_b.cell(3, 1, "BANCO").font = BOLD
    ws_b.cell(3, 2, "QUANTIDADE").font = BOLD
    ws_b.cell(3, 3, "VALOR LIBERADO (R$)").font = BOLD
    bd = por_banco(pago_falta)
    for i, (banco, d) in enumerate(sorted(bd.items(), key=lambda x: -x[1]['val']), 4):
        ws_b.cell(i, 1, banco)
        ws_b.cell(i, 2, d['qtd'])
        ws_b.cell(i, 3, d['val']).number_format = 'R$ #,##0.00'
    tr2 = len(bd) + 4
    ws_b.cell(tr2, 1, "TOTAL").font = BOLD
    ws_b.cell(tr2, 2, sum(d['qtd'] for d in bd.values())).font = BOLD
    c = ws_b.cell(tr2, 3, soma(pago_falta)); c.font = BOLD; c.number_format = 'R$ #,##0.00'
    ws_b.column_dimensions['A'].width = 28
    ws_b.column_dimensions['B'].width = 14
    ws_b.column_dimensions['C'].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ── rotas ──────────────────────────────────────────────────
@app.route('/')
def index():
    with open(os.path.join(os.path.dirname(__file__), 'static', 'index.html'), encoding='utf-8') as f:
        return f.read()

@app.route('/processar', methods=['POST'])
def processar_route():
    if 'japa' not in request.files or 'nossa' not in request.files:
        return jsonify({'erro': 'Envie as duas planilhas'}), 400

    sid = str(uuid.uuid4())[:8]
    path_japa  = os.path.join(UPLOAD_FOLDER,  f'{sid}_japa.xlsx')
    path_nossa = os.path.join(UPLOAD_FOLDER,  f'{sid}_nossa.xlsx')
    path_res   = os.path.join(RESULT_FOLDER,  f'{sid}_resultado.xlsx')

    request.files['japa'].save(path_japa)
    request.files['nossa'].save(path_nossa)

    try:
        resultado = processar(path_japa, path_nossa)
        buf = gerar_xlsx(resultado)
        with open(path_res, 'wb') as f:
            f.write(buf.read())

        # remover _listas e _header antes de retornar JSON
        resultado.pop('_listas')
        resultado.pop('_header')
        resultado['sid'] = sid
        return jsonify(resultado)
    except Exception as e:
        return jsonify({'erro': str(e)}), 500

@app.route('/download/<sid>')
def download(sid):
    path = os.path.join(RESULT_FOLDER, f'{sid}_resultado.xlsx')
    if not os.path.exists(path):
        return 'Arquivo não encontrado', 404
    return send_file(path, as_attachment=True, download_name='RESULTADO_ANALISE.xlsx')

if __name__ == '__main__':
    app.run(debug=True, port=5000)
